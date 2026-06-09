#!/usr/bin/env python3
"""
JAMF Classic API - Policy Export Script
Fetches all policies with payloads (packages, scripts, DMGs, inventory, etc.)
and their scopes using OAuth client credentials.

Token is automatically refreshed when it is close to expiry or when a 401
is received mid-run (server-side invalidation / clock skew).
"""

import argparse
import csv
import datetime
import io
import json
import sys
import time
import requests
from typing import Optional


# ---------------------------------------------------------------------------
# Auth — token manager with expiry tracking and auto-refresh
# ---------------------------------------------------------------------------

# Refresh when this fraction of the token lifetime remains.
# e.g. 0.1 = refresh in the last 10% of the token's life.
# This adapts automatically to short-lived tokens (like JAMF's 59s default).
_TOKEN_REFRESH_FRACTION = 0.10

# Hard floor: always refresh if fewer than this many seconds remain,
# regardless of the fraction calculation.
_TOKEN_REFRESH_FLOOR_SECS = 5


class TokenManager:
    """
    Holds an OAuth Bearer token and transparently re-authenticates when:
      • the token is in the last _TOKEN_REFRESH_FRACTION of its lifetime, OR
      • a request comes back 401 (server-side invalidation / clock skew).

    Using a fraction (rather than a fixed buffer) means a 59-second token
    refreshes only in its last ~6 seconds, instead of on every single request.
    """

    def __init__(self, jamf_url: str, client_id: str, client_secret: str) -> None:
        self._jamf_url = jamf_url
        self._client_id = client_id
        self._client_secret = client_secret
        self._token: Optional[str] = None
        self._expires_at: float = 0.0   # monotonic seconds
        self._expires_in: int = 0       # original lifetime from server

    def _refresh_threshold(self) -> float:
        """Seconds before expiry at which we proactively refresh."""
        return max(
            _TOKEN_REFRESH_FLOOR_SECS,
            self._expires_in * _TOKEN_REFRESH_FRACTION,
        )

    def _fetch(self, reason: str = "initial") -> None:
        token_url = f"{self._jamf_url}/api/oauth/token"
        payload = {
            "grant_type": "client_credentials",
            "client_id": self._client_id,
            "client_secret": self._client_secret,
        }
        resp = requests.post(token_url, data=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        self._token = data["access_token"]
        self._expires_in = int(data.get("expires_in", 1800))
        self._expires_at = time.monotonic() + self._expires_in
        threshold = self._refresh_threshold()
        print(
            f"[auth] Token acquired ({reason}) — "
            f"valid for {self._expires_in}s, "
            f"will refresh with ~{threshold:.0f}s remaining",
            file=sys.stderr,
        )

    @property
    def token(self) -> str:
        """Return a valid token, refreshing proactively when near expiry."""
        remaining = self._expires_at - time.monotonic()
        if self._token is None or remaining < self._refresh_threshold():
            if self._token is not None:
                print(
                    f"[auth] Token expiring soon ({remaining:.0f}s left) — refreshing …",
                    file=sys.stderr,
                )
                self._fetch(reason="proactive refresh")
            else:
                self._fetch(reason="initial")
        return self._token  # type: ignore[return-value]

    def force_refresh(self) -> None:
        """Call this after a 401 to obtain a brand-new token immediately."""
        print("[auth] 401 received — forcing token refresh …", file=sys.stderr)
        self._token = None
        self._expires_at = 0.0
        self._fetch(reason="forced after 401")


# ---------------------------------------------------------------------------
# Classic API helpers
# ---------------------------------------------------------------------------

def classic_get(url: str, tm: TokenManager, retries: int = 2) -> dict:
    """
    GET a Classic API endpoint and return parsed JSON.
    Automatically refreshes the token and retries up to `retries` times on 401.
    """
    for attempt in range(1, retries + 2):   # e.g. retries=2 → up to 3 attempts
        headers = {
            "Authorization": f"Bearer {tm.token}",
            "Accept": "application/json",
        }
        resp = requests.get(url, headers=headers, timeout=60)

        if resp.status_code == 401 and attempt <= retries:
            tm.force_refresh()
            continue    # retry with fresh token

        resp.raise_for_status()
        return resp.json()

    raise RuntimeError("classic_get: exhausted retries without success")


def get_all_policy_ids(jamf_url: str, tm: TokenManager) -> list[dict]:
    """Return a list of {id, name} for every policy."""
    data = classic_get(f"{jamf_url}/JSSResource/policies", tm)
    return data.get("policies", [])


def get_policy_detail(jamf_url: str, tm: TokenManager, policy_id: int) -> dict:
    """Return the full detail record for a single policy."""
    data = classic_get(f"{jamf_url}/JSSResource/policies/id/{policy_id}", tm)
    return data.get("policy", {})


# ---------------------------------------------------------------------------
# Payload extraction helpers
# ---------------------------------------------------------------------------
#
# The JAMF Classic API is wildly inconsistent across versions and policy age.
# The SAME field can appear as:
#   - A bare list:            "scripts": [{...}, {...}]          (modern JSON)
#   - A wrapped dict+list:    "scripts": {"script": [{...}]}     (older XML→JSON)
#   - A wrapped dict+single:  "scripts": {"script": {...}}       (single item, old)
#   - A size-only dict:       "scripts": {"size": "0"}           (empty, old)
#   - An empty list:          "scripts": []                      (empty, modern)
#   - An empty dict:          "scripts": {}                      (empty, alt)
#   - A list with empty str:  "printers": [""]                   (JAMF quirk)
#
# _coerce_list() normalises ALL of these into a clean list of dicts.

def _safe_get(obj, *keys, default=None):
    cur = obj
    for key in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key, default)
        if cur is None:
            return default
    return cur


def _coerce_list(val, inner_key: str = None) -> list:
    """
    Universal normaliser. Given a value that may be any of the shapes above,
    return a clean list of dicts.

    If inner_key is provided (e.g. "script"), check for the old wrapped shape
    {"script": ...} before treating the dict as a single item.
    """
    if not val:
        return []

    # Bare list (modern API shape)
    if isinstance(val, list):
        return [i for i in val if isinstance(i, dict) and i]

    if isinstance(val, dict):
        # Size-only sentinel → empty
        without_size = {k: v for k, v in val.items() if k != "size"}
        if not without_size:
            return []

        # Old wrapped shape: {"script": <list-or-dict>}
        if inner_key and inner_key in val:
            inner = val[inner_key]
            if isinstance(inner, list):
                return [i for i in inner if isinstance(i, dict) and i]
            if isinstance(inner, dict) and inner:
                return [inner]
            return []

        # Plain single-item dict
        return [without_size]

    return []


def extract_packages(policy: dict) -> list[dict]:
    pkg_cfg = policy.get("package_configuration") or {}
    raw = pkg_cfg.get("packages", [])
    # modern: packages is a bare list
    # old:    packages is {"package": [...]} or {"size":"0"}
    pkgs = _coerce_list(raw, inner_key="package")
    return [
        {
            "id": p.get("id"),
            "name": p.get("name"),
            "action": p.get("action"),
            "fut": p.get("fut"),
            "feu": p.get("feu"),
        }
        for p in pkgs
        if p.get("name")
    ]


def extract_scripts(policy: dict) -> list[dict]:
    raw = policy.get("scripts", [])
    # modern: bare list of script dicts
    # old:    {"script": [...]} or {"size":"0"}
    scripts = _coerce_list(raw, inner_key="script")
    return [
        {
            "id": s.get("id"),
            "name": s.get("name"),
            "priority": s.get("priority"),
            "parameter4":  s.get("parameter4")  or "",
            "parameter5":  s.get("parameter5")  or "",
            "parameter6":  s.get("parameter6")  or "",
            "parameter7":  s.get("parameter7")  or "",
            "parameter8":  s.get("parameter8")  or "",
            "parameter9":  s.get("parameter9")  or "",
            "parameter10": s.get("parameter10") or "",
            "parameter11": s.get("parameter11") or "",
        }
        for s in scripts
        if s.get("name")
    ]


def extract_printers(policy: dict) -> list[dict]:
    raw = policy.get("printers", [])
    printers = _coerce_list(raw, inner_key="printer")
    return [
        {"id": p.get("id"), "name": p.get("name"), "action": p.get("action")}
        for p in printers
        if p.get("name")
    ]


def extract_dock_items(policy: dict) -> list[dict]:
    raw = policy.get("dock_items", [])
    items = _coerce_list(raw, inner_key="dock_item")
    return [
        {"id": d.get("id"), "name": d.get("name"), "action": d.get("action")}
        for d in items
        if d.get("name")
    ]


def extract_account_maintenance(policy: dict) -> dict:
    return policy.get("account_maintenance", {}) or {}


def extract_maintenance(policy: dict) -> dict:
    return policy.get("maintenance", {}) or {}


def extract_files_processes(policy: dict) -> dict:
    return policy.get("files_processes", {}) or {}


def extract_disk_encryption(policy: dict) -> dict:
    return policy.get("disk_encryption", {}) or {}


def extract_reboot(policy: dict) -> dict:
    return policy.get("reboot", {}) or {}


def extract_inventory(policy: dict) -> dict:
    """Inventory Collection and Inventory Preload settings."""
    return {
        "recon": policy.get("recon", {}) or {},
        "inventory_preload": policy.get("inventory_preload", {}) or {},
    }


def extract_self_service(policy: dict) -> dict:
    return policy.get("self_service", {}) or {}


def extract_triggers(policy: dict) -> dict:
    general = policy.get("general", {}) or {}
    return {
        "trigger": general.get("trigger"),
        "trigger_checkin": general.get("trigger_checkin"),
        "trigger_enrollment_complete": general.get("trigger_enrollment_complete"),
        "trigger_login": general.get("trigger_login"),
        "trigger_logout": general.get("trigger_logout"),
        "trigger_network_state_changed": general.get("trigger_network_state_changed"),
        "trigger_startup": general.get("trigger_startup"),
        "trigger_other": general.get("trigger_other"),
    }


# ---------------------------------------------------------------------------
# Scope extraction
# ---------------------------------------------------------------------------

def extract_scope(policy: dict) -> dict:
    scope = policy.get("scope", {}) or {}

    def grab(container, key: str, inner_key: str) -> list:
        """Pull a scope collection — handles both bare list and old wrapped dict."""
        if not isinstance(container, dict):
            return []
        return _coerce_list(container.get(key, []), inner_key=inner_key)

    lim  = scope.get("limitations") or {}
    excl = scope.get("exclusions")  or {}

    return {
        "all_computers": scope.get("all_computers", False),
        "all_jss_users": scope.get("all_jss_users", False),
        # Targets
        "computers":      grab(scope, "computers",      "computer"),
        "computer_groups":grab(scope, "computer_groups","computer_group"),
        "buildings":      grab(scope, "buildings",      "building"),
        "departments":    grab(scope, "departments",    "department"),
        "jss_users":      grab(scope, "jss_users",      "jss_user"),
        "jss_user_groups":grab(scope, "jss_user_groups","jss_user_group"),
        # Limitations
        "limitations": {
            "users":            grab(lim, "users",            "user"),
            "user_groups":      grab(lim, "user_groups",      "user_group"),
            "network_segments": grab(lim, "network_segments", "network_segment"),
            "ibeacons":         grab(lim, "ibeacons",         "ibeacon"),
        },
        # Exclusions
        "exclusions": {
            "computers":        grab(excl, "computers",        "computer"),
            "computer_groups":  grab(excl, "computer_groups",  "computer_group"),
            "buildings":        grab(excl, "buildings",        "building"),
            "departments":      grab(excl, "departments",      "department"),
            "users":            grab(excl, "users",            "user"),
            "user_groups":      grab(excl, "user_groups",      "user_group"),
            "network_segments": grab(excl, "network_segments", "network_segment"),
            "ibeacons":         grab(excl, "ibeacons",         "ibeacon"),
        },
    }


# ---------------------------------------------------------------------------
# Main aggregation
# ---------------------------------------------------------------------------

def build_policy_record(policy: dict) -> dict:
    general = policy.get("general", {}) or {}
    return {
        "id": general.get("id"),
        "name": general.get("name"),
        "enabled": general.get("enabled"),
        "category": _safe_get(general, "category", "name"),
        "frequency": general.get("frequency"),
        "triggers": extract_triggers(policy),
        "payloads": {
            "packages": extract_packages(policy),
            "scripts": extract_scripts(policy),
            "printers": extract_printers(policy),
            "dock_items": extract_dock_items(policy),
            "account_maintenance": extract_account_maintenance(policy),
            "maintenance": extract_maintenance(policy),
            "files_processes": extract_files_processes(policy),
            "disk_encryption": extract_disk_encryption(policy),
            "reboot": extract_reboot(policy),
            "inventory": extract_inventory(policy),
        },
        "self_service": extract_self_service(policy),
        "scope": extract_scope(policy),
    }


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def _stamp(path: Optional[str]) -> Optional[str]:
    """Insert a YYYYMMDD_HHMMSS timestamp before the file extension."""
    if not path:
        return path
    import os
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    return f"{base}_{ts}{ext}"


def write_json(records: list[dict], output_file: Optional[str]) -> None:
    out = json.dumps(records, indent=2, default=str)
    output_file = _stamp(output_file)
    if output_file:
        with open(output_file, "w") as f:
            f.write(out)
        print(f"[✓] Written {len(records)} policies → {output_file}", file=sys.stderr)
    else:
        print(out)


def write_text_summary(records: list[dict], output_file: Optional[str]) -> None:
    lines = []
    sep = "=" * 72

    for rec in records:
        lines.append(sep)
        lines.append(f"Policy: {rec['name']}  (ID: {rec['id']})")
        lines.append(f"  Enabled  : {rec['enabled']}")
        lines.append(f"  Category : {rec['category']}")
        lines.append(f"  Frequency: {rec['frequency']}")

        t = rec["triggers"]
        active = [k for k, v in t.items() if v and v not in (False, "EVENT_NONE", "")]
        lines.append(f"  Triggers : {', '.join(active) if active else 'none'}")

        p = rec["payloads"]

        if p["packages"]:
            lines.append("  Packages:")
            for pkg in p["packages"]:
                lines.append(f"    • {pkg['name']}  action={pkg['action']}")

        if p["scripts"]:
            lines.append("  Scripts:")
            for s in p["scripts"]:
                params = {k: v for k, v in s.items() if k.startswith("parameter") and v}
                param_str = "  params=" + json.dumps(params) if params else ""
                lines.append(f"    • {s['name']}  priority={s['priority']}{param_str}")

        if p["printers"]:
            lines.append("  Printers:")
            for pr in p["printers"]:
                lines.append(f"    • {pr['name']}  action={pr['action']}")

        if p["dock_items"]:
            lines.append("  Dock Items:")
            for di in p["dock_items"]:
                lines.append(f"    • {di['name']}  action={di['action']}")

        maint = p["maintenance"]
        if maint and any(maint.values()):
            lines.append(f"  Maintenance: {json.dumps(maint)}")

        fp = p["files_processes"]
        if fp and any(fp.values()):
            lines.append(f"  Files/Processes: {json.dumps(fp)}")

        de = p["disk_encryption"]
        if de and any(de.values()):
            lines.append(f"  Disk Encryption: {json.dumps(de)}")

        inv = p["inventory"]
        recon = inv.get("recon", {})
        if recon and recon.get("recon"):
            lines.append(f"  Inventory (Recon): {json.dumps(recon)}")

        sc = rec["scope"]
        lines.append("  Scope:")
        if sc["all_computers"]:
            lines.append("    Targets: All Computers")
        else:
            for c in sc["computers"]:
                lines.append(f"    Computer      : {c.get('name')} (id={c.get('id')})")
            for cg in sc["computer_groups"]:
                lines.append(f"    Computer Group: {cg.get('name')} (id={cg.get('id')})")
            for b in sc["buildings"]:
                lines.append(f"    Building      : {b.get('name')}")
            for d in sc["departments"]:
                lines.append(f"    Department    : {d.get('name')}")

        excl = sc["exclusions"]
        excl_items = (
            excl["computers"] + excl["computer_groups"] +
            excl["buildings"] + excl["departments"]
        )
        if excl_items:
            lines.append("    Exclusions:")
            for item in excl_items:
                lines.append(f"      - {item.get('name')}")

        lim = sc["limitations"]
        lim_items = lim["users"] + lim["user_groups"] + lim["network_segments"]
        if lim_items:
            lines.append("    Limitations:")
            for item in lim_items:
                lines.append(f"      - {item.get('name')}")

    lines.append(sep)
    output = "\n".join(lines)

    output_file = _stamp(output_file)
    if output_file:
        with open(output_file, "w") as f:
            f.write(output)
        print(f"[✓] Written {len(records)} policies → {output_file}", file=sys.stderr)
    else:
        print(output)


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def _names(items: list) -> str:
    """Join a list of {name, …} dicts into a semicolon-separated string."""
    return "; ".join(i.get("name", "") for i in items if isinstance(i, dict) and i.get("name"))


def _scope_summary(sc: dict) -> str:
    """One-line scope summary: all computers, or list of targets."""
    if sc.get("all_computers"):
        return "All Computers"
    parts = []
    for c in sc.get("computers", []):
        parts.append(c.get("name", ""))
    for g in sc.get("computer_groups", []):
        parts.append(f"[Group] {g.get('name', '')}")
    for b in sc.get("buildings", []):
        parts.append(f"[Building] {b.get('name', '')}")
    for d in sc.get("departments", []):
        parts.append(f"[Dept] {d.get('name', '')}")
    return "; ".join(p for p in parts if p)


def write_csv(records: list[dict], output_file: Optional[str]) -> None:
    """
    Write THREE CSV files (or three sections to stdout):
      1. policies.csv        — one row per policy, all general/scope/maintenance fields
      2. policies_pkgs.csv   — one row per package payload (policy_id, policy_name, pkg fields)
      3. policies_scripts.csv— one row per script payload (policy_id, policy_name, script + all params)

    When output_file is given, e.g. "policies.csv", the other two are written as
    "policies_pkgs.csv" and "policies_scripts.csv" alongside it.
    """

    # --- Sheet 1: Policies (one row per policy) ---
    policy_fields = [
        "policy_id", "policy_name", "enabled", "category", "frequency",
        # Triggers
        "trigger", "trigger_checkin", "trigger_enrollment_complete",
        "trigger_login", "trigger_logout", "trigger_network_state_changed",
        "trigger_startup", "trigger_other",
        # Payload counts
        "package_count", "script_count", "printer_count", "dock_item_count",
        # Inline payloads (semicolon-separated names for quick scan)
        "packages", "scripts", "printers", "dock_items",
        # Maintenance
        "maint_recon", "maint_reset_name", "maint_install_cached_pkgs",
        "maint_heal", "maint_permissions", "maint_byhost",
        "maint_system_cache", "maint_user_cache", "maint_verify",
        # Files & Processes
        "fp_search_path", "fp_delete_file", "fp_search_process",
        "fp_kill_process", "fp_run_command",
        # Disk encryption
        "disk_encryption_action",
        # Reboot
        "reboot_no_user", "reboot_user_logged_in", "reboot_minutes",
        # Inventory
        "inventory_recon",
        # Scope
        "scope_all_computers", "scope_targets",
        "scope_computers", "scope_computer_groups",
        "scope_buildings", "scope_departments",
        "scope_limit_users", "scope_limit_user_groups", "scope_limit_network_segments",
        "scope_excl_computers", "scope_excl_groups",
        "scope_excl_buildings", "scope_excl_departments",
        # Self service
        "self_service_enabled", "self_service_name",
        "self_service_description", "self_service_button_text",
        "self_service_notification",
    ]

    # --- Sheet 2: Packages ---
    pkg_fields = [
        "policy_id", "policy_name", "enabled", "category",
        "pkg_id", "pkg_name", "action", "fill_user_template", "fill_existing_users",
    ]

    # --- Sheet 3: Scripts ---
    script_fields = [
        "policy_id", "policy_name", "enabled", "category",
        "script_id", "script_name", "priority",
        "parameter4", "parameter5", "parameter6", "parameter7",
        "parameter8", "parameter9", "parameter10", "parameter11",
    ]

    policy_rows, pkg_rows, script_rows = [], [], []

    for rec in records:
        p   = rec["payloads"]
        t   = rec["triggers"]
        sc  = rec["scope"]
        m   = p.get("maintenance") or {}
        fp  = p.get("files_processes") or {}
        de  = p.get("disk_encryption") or {}
        rb  = p.get("reboot") or {}
        inv = p.get("inventory") or {}
        recon = inv.get("recon") or {}
        ss  = rec.get("self_service") or {}
        pid = rec["id"]
        pname = rec["name"]

        policy_rows.append({
            "policy_id": pid,
            "policy_name": pname,
            "enabled": rec["enabled"],
            "category": rec["category"],
            "frequency": rec["frequency"],
            "trigger": t.get("trigger"),
            "trigger_checkin": t.get("trigger_checkin"),
            "trigger_enrollment_complete": t.get("trigger_enrollment_complete"),
            "trigger_login": t.get("trigger_login"),
            "trigger_logout": t.get("trigger_logout"),
            "trigger_network_state_changed": t.get("trigger_network_state_changed"),
            "trigger_startup": t.get("trigger_startup"),
            "trigger_other": t.get("trigger_other"),
            "package_count": len(p.get("packages", [])),
            "script_count": len(p.get("scripts", [])),
            "printer_count": len(p.get("printers", [])),
            "dock_item_count": len(p.get("dock_items", [])),
            "packages":   "; ".join(f"{x['name']} ({x['action']})" for x in p.get("packages", []) if x.get("name")),
            "scripts":    "; ".join(f"{x['name']} [{x['priority']}]" for x in p.get("scripts", []) if x.get("name")),
            "printers":   _names(p.get("printers", [])),
            "dock_items": _names(p.get("dock_items", [])),
            "maint_recon": m.get("recon"),
            "maint_reset_name": m.get("reset_name"),
            "maint_install_cached_pkgs": m.get("install_all_cached_packages"),
            "maint_heal": m.get("heal"),
            "maint_permissions": m.get("permissions"),
            "maint_byhost": m.get("byhost"),
            "maint_system_cache": m.get("system_cache"),
            "maint_user_cache": m.get("user_cache"),
            "maint_verify": m.get("verify"),
            "fp_search_path": fp.get("search_by_path"),
            "fp_delete_file": fp.get("delete_file"),
            "fp_search_process": fp.get("search_for_process"),
            "fp_kill_process": fp.get("kill_process"),
            "fp_run_command": fp.get("run_command"),
            "disk_encryption_action": de.get("action"),
            "reboot_no_user": rb.get("no_user_logged_in"),
            "reboot_user_logged_in": rb.get("user_logged_in"),
            "reboot_minutes": rb.get("minutes_until_reboot"),
            "inventory_recon": recon.get("recon"),
            "scope_all_computers": sc.get("all_computers"),
            "scope_targets": _scope_summary(sc),
            "scope_computers": _names(sc.get("computers", [])),
            "scope_computer_groups": _names(sc.get("computer_groups", [])),
            "scope_buildings": _names(sc.get("buildings", [])),
            "scope_departments": _names(sc.get("departments", [])),
            "scope_limit_users": _names(sc.get("limitations", {}).get("users", [])),
            "scope_limit_user_groups": _names(sc.get("limitations", {}).get("user_groups", [])),
            "scope_limit_network_segments": _names(sc.get("limitations", {}).get("network_segments", [])),
            "scope_excl_computers": _names(sc.get("exclusions", {}).get("computers", [])),
            "scope_excl_groups": _names(sc.get("exclusions", {}).get("computer_groups", [])),
            "scope_excl_buildings": _names(sc.get("exclusions", {}).get("buildings", [])),
            "scope_excl_departments": _names(sc.get("exclusions", {}).get("departments", [])),
            "self_service_enabled": ss.get("use_for_self_service"),
            "self_service_name": ss.get("self_service_display_name"),
            "self_service_description": ss.get("self_service_description"),
            "self_service_button_text": ss.get("install_button_text"),
            "self_service_notification": ss.get("notification"),
        })

        for pkg in p.get("packages", []):
            if not pkg.get("name"):
                continue
            pkg_rows.append({
                "policy_id": pid,
                "policy_name": pname,
                "enabled": rec["enabled"],
                "category": rec["category"],
                "pkg_id": pkg.get("id"),
                "pkg_name": pkg.get("name"),
                "action": pkg.get("action"),
                "fill_user_template": pkg.get("fut"),
                "fill_existing_users": pkg.get("feu"),
            })

        for s in p.get("scripts", []):
            if not s.get("name"):
                continue
            script_rows.append({
                "policy_id": pid,
                "policy_name": pname,
                "enabled": rec["enabled"],
                "category": rec["category"],
                "script_id": s.get("id"),
                "script_name": s.get("name"),
                "priority": s.get("priority"),
                "parameter4":  s.get("parameter4")  or "",
                "parameter5":  s.get("parameter5")  or "",
                "parameter6":  s.get("parameter6")  or "",
                "parameter7":  s.get("parameter7")  or "",
                "parameter8":  s.get("parameter8")  or "",
                "parameter9":  s.get("parameter9")  or "",
                "parameter10": s.get("parameter10") or "",
                "parameter11": s.get("parameter11") or "",
            })

    def _write(rows, fields, path, label):
        if path:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields)
                w.writeheader()
                w.writerows(rows)
            print(f"[✓] Written {len(rows)} rows → {path}  ({label})", file=sys.stderr)
        else:
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=fields)
            w.writeheader()
            w.writerows(rows)
            print(f"\n### {label} ###")
            print(buf.getvalue())

    # Derive sibling filenames from the base output path, with timestamp
    if output_file:
        import os
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_file)
        # Insert timestamp before extension: policies.csv → policies_20240506_143022.csv
        output_file = f"{base}_{ts}{ext}"
        pkg_path    = f"{base}_packages_{ts}{ext}"
        script_path = f"{base}_scripts_{ts}{ext}"
    else:
        pkg_path = script_path = None

    _write(policy_rows, policy_fields, output_file, "policies")
    _write(pkg_rows,    pkg_fields,    pkg_path,    "packages")
    _write(script_rows, script_fields, script_path, "scripts")



# ---------------------------------------------------------------------------
# XLSX workbook output — one tab per data type, formatted
# ---------------------------------------------------------------------------

def write_xlsx(records: list[dict], output_file: Optional[str]) -> None:
    """
    Write a single timestamped .xlsx workbook with three tabs:
      • Policies  — one row per policy (general + triggers + payloads summary + scope + self service)
      • Packages  — one row per package payload
      • Scripts   — one row per script payload with all parameters
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        print(
            "[!] openpyxl is required for xlsx output but is not installed.\n"
            "    Install it with:\n"
            "        pip install openpyxl\n"
            "    Then re-run the script.",
            file=sys.stderr,
        )
        sys.exit(1)

    # ── colour palette ───────────────────────────────────────────────────────
    HDR_BG  = "1F4E79"
    HDR_FG  = "FFFFFF"
    ALT_BG  = "D6E4F0"
    TAB_COLS = {
        "Policies": "1F4E79",
        "Packages": "2E75B6",
        "Scripts":  "2F5496",
    }

    thin        = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers):
        hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=10)
        hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def write_rows(ws, headers, rows):
        alt_fill   = PatternFill("solid", fgColor=ALT_BG)
        body_font  = Font(name="Arial", size=9)
        body_align = Alignment(vertical="top", wrap_text=False)
        for r_idx, row in enumerate(rows, start=2):
            fill = alt_fill if r_idx % 2 == 0 else None
            for c_idx, key in enumerate(headers, start=1):
                val = row.get(key, "")
                if val is None:
                    val = ""
                elif isinstance(val, bool):
                    val = "Yes" if val else "No"
                elif not isinstance(val, (int, float)):
                    val = str(val)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = body_font
                cell.alignment = body_align
                cell.border    = cell_border
                if fill:
                    cell.fill = fill

    def autofit(ws, headers, rows, min_w=10, max_w=60):
        for c_idx, key in enumerate(headers, start=1):
            vals = [str(key)] + [str(r.get(key, "") or "") for r in rows[:200]]
            best = min(max(max(len(v.split("\n")[0]) for v in vals) + 2, min_w), max_w)
            ws.column_dimensions[get_column_letter(c_idx)].width = best

    def add_sheet(wb, name, headers, rows):
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = TAB_COLS.get(name, "4472C4")
        style_header(ws, headers)
        write_rows(ws, headers, rows)
        autofit(ws, headers, rows)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── header / key definitions ─────────────────────────────────────────────
    policy_headers = [
        "Policy ID", "Policy Name", "Enabled", "Category", "Frequency",
        "Trigger", "Trigger: Checkin", "Trigger: Enrollment",
        "Trigger: Login", "Trigger: Logout", "Trigger: Network Change",
        "Trigger: Startup", "Trigger: Other",
        "Package Count", "Script Count", "Printer Count", "Dock Item Count",
        "Packages", "Scripts", "Printers", "Dock Items",
        "Maint: Recon", "Maint: Reset Name", "Maint: Install Cached Pkgs",
        "Maint: Heal", "Maint: Permissions", "Maint: ByHost",
        "Maint: System Cache", "Maint: User Cache", "Maint: Verify",
        "Files: Search Path", "Files: Delete", "Files: Search Process",
        "Files: Kill Process", "Files: Run Command",
        "Disk Encryption Action",
        "Reboot: No User", "Reboot: User Logged In", "Reboot: Minutes",
        "Inventory Recon",
        "Scope: All Computers", "Scope: Targets",
        "Scope: Computers", "Scope: Computer Groups",
        "Scope: Buildings", "Scope: Departments",
        "Limit: Users", "Limit: User Groups", "Limit: Network Segments",
        "Excl: Computers", "Excl: Computer Groups",
        "Excl: Buildings", "Excl: Departments",
        "Self Service Enabled", "Self Service Name",
        "Self Service Description", "Self Service Button",
        "Self Service Notification",
    ]
    policy_keys = [
        "policy_id", "policy_name", "enabled", "category", "frequency",
        "trigger", "trigger_checkin", "trigger_enrollment_complete",
        "trigger_login", "trigger_logout", "trigger_network_state_changed",
        "trigger_startup", "trigger_other",
        "package_count", "script_count", "printer_count", "dock_item_count",
        "packages", "scripts", "printers", "dock_items",
        "maint_recon", "maint_reset_name", "maint_install_cached_pkgs",
        "maint_heal", "maint_permissions", "maint_byhost",
        "maint_system_cache", "maint_user_cache", "maint_verify",
        "fp_search_path", "fp_delete_file", "fp_search_process",
        "fp_kill_process", "fp_run_command",
        "disk_encryption_action",
        "reboot_no_user", "reboot_user_logged_in", "reboot_minutes",
        "inventory_recon",
        "scope_all_computers", "scope_targets",
        "scope_computers", "scope_computer_groups",
        "scope_buildings", "scope_departments",
        "scope_limit_users", "scope_limit_user_groups", "scope_limit_network_segments",
        "scope_excl_computers", "scope_excl_groups",
        "scope_excl_buildings", "scope_excl_departments",
        "self_service_enabled", "self_service_name",
        "self_service_description", "self_service_button_text",
        "self_service_notification",
    ]

    pkg_headers = [
        "Policy ID", "Policy Name", "Enabled", "Category",
        "Package ID", "Package Name", "Action",
        "Fill User Template", "Fill Existing Users",
    ]
    pkg_keys = [
        "policy_id", "policy_name", "enabled", "category",
        "pkg_id", "pkg_name", "action", "fill_user_template", "fill_existing_users",
    ]

    script_headers = [
        "Policy ID", "Policy Name", "Enabled", "Category",
        "Script ID", "Script Name", "Priority",
        "Parameter 4", "Parameter 5", "Parameter 6", "Parameter 7",
        "Parameter 8", "Parameter 9", "Parameter 10", "Parameter 11",
    ]
    script_keys = [
        "policy_id", "policy_name", "enabled", "category",
        "script_id", "script_name", "priority",
        "parameter4", "parameter5", "parameter6", "parameter7",
        "parameter8", "parameter9", "parameter10", "parameter11",
    ]

    # ── build row data ───────────────────────────────────────────────────────
    policy_rows, pkg_rows, script_rows = [], [], []

    for rec in records:
        p     = rec["payloads"]
        t     = rec["triggers"]
        sc    = rec["scope"]
        m     = p.get("maintenance") or {}
        fp    = p.get("files_processes") or {}
        de    = p.get("disk_encryption") or {}
        rb    = p.get("reboot") or {}
        inv   = p.get("inventory") or {}
        recon = inv.get("recon") or {}
        ss    = rec.get("self_service") or {}
        pid   = rec["id"]
        pname = rec["name"]

        policy_rows.append({
            "policy_id": pid, "policy_name": pname,
            "enabled": rec["enabled"], "category": rec["category"],
            "frequency": rec["frequency"],
            "trigger": t.get("trigger"),
            "trigger_checkin": t.get("trigger_checkin"),
            "trigger_enrollment_complete": t.get("trigger_enrollment_complete"),
            "trigger_login": t.get("trigger_login"),
            "trigger_logout": t.get("trigger_logout"),
            "trigger_network_state_changed": t.get("trigger_network_state_changed"),
            "trigger_startup": t.get("trigger_startup"),
            "trigger_other": t.get("trigger_other"),
            "package_count": len(p.get("packages", [])),
            "script_count":  len(p.get("scripts", [])),
            "printer_count": len(p.get("printers", [])),
            "dock_item_count": len(p.get("dock_items", [])),
            "packages":   "; ".join(f"{x['name']} ({x['action']})" for x in p.get("packages", []) if x.get("name")),
            "scripts":    "; ".join(f"{x['name']} [{x['priority']}]" for x in p.get("scripts", []) if x.get("name")),
            "printers":   _names(p.get("printers", [])),
            "dock_items": _names(p.get("dock_items", [])),
            "maint_recon": m.get("recon"),
            "maint_reset_name": m.get("reset_name"),
            "maint_install_cached_pkgs": m.get("install_all_cached_packages"),
            "maint_heal": m.get("heal"),
            "maint_permissions": m.get("permissions"),
            "maint_byhost": m.get("byhost"),
            "maint_system_cache": m.get("system_cache"),
            "maint_user_cache": m.get("user_cache"),
            "maint_verify": m.get("verify"),
            "fp_search_path": fp.get("search_by_path"),
            "fp_delete_file": fp.get("delete_file"),
            "fp_search_process": fp.get("search_for_process"),
            "fp_kill_process": fp.get("kill_process"),
            "fp_run_command": fp.get("run_command"),
            "disk_encryption_action": de.get("action"),
            "reboot_no_user": rb.get("no_user_logged_in"),
            "reboot_user_logged_in": rb.get("user_logged_in"),
            "reboot_minutes": rb.get("minutes_until_reboot"),
            "inventory_recon": recon.get("recon"),
            "scope_all_computers": sc.get("all_computers"),
            "scope_targets": _scope_summary(sc),
            "scope_computers": _names(sc.get("computers", [])),
            "scope_computer_groups": _names(sc.get("computer_groups", [])),
            "scope_buildings": _names(sc.get("buildings", [])),
            "scope_departments": _names(sc.get("departments", [])),
            "scope_limit_users": _names(sc.get("limitations", {}).get("users", [])),
            "scope_limit_user_groups": _names(sc.get("limitations", {}).get("user_groups", [])),
            "scope_limit_network_segments": _names(sc.get("limitations", {}).get("network_segments", [])),
            "scope_excl_computers": _names(sc.get("exclusions", {}).get("computers", [])),
            "scope_excl_groups": _names(sc.get("exclusions", {}).get("computer_groups", [])),
            "scope_excl_buildings": _names(sc.get("exclusions", {}).get("buildings", [])),
            "scope_excl_departments": _names(sc.get("exclusions", {}).get("departments", [])),
            "self_service_enabled": ss.get("use_for_self_service"),
            "self_service_name": ss.get("self_service_display_name"),
            "self_service_description": ss.get("self_service_description"),
            "self_service_button_text": ss.get("install_button_text"),
            "self_service_notification": ss.get("notification"),
        })

        for pkg in p.get("packages", []):
            if not pkg.get("name"):
                continue
            pkg_rows.append({
                "policy_id": pid, "policy_name": pname,
                "enabled": rec["enabled"], "category": rec["category"],
                "pkg_id": pkg.get("id"), "pkg_name": pkg.get("name"),
                "action": pkg.get("action"),
                "fill_user_template": pkg.get("fut"),
                "fill_existing_users": pkg.get("feu"),
            })

        for s in p.get("scripts", []):
            if not s.get("name"):
                continue
            script_rows.append({
                "policy_id": pid, "policy_name": pname,
                "enabled": rec["enabled"], "category": rec["category"],
                "script_id": s.get("id"), "script_name": s.get("name"),
                "priority": s.get("priority"),
                "parameter4":  s.get("parameter4")  or "",
                "parameter5":  s.get("parameter5")  or "",
                "parameter6":  s.get("parameter6")  or "",
                "parameter7":  s.get("parameter7")  or "",
                "parameter8":  s.get("parameter8")  or "",
                "parameter9":  s.get("parameter9")  or "",
                "parameter10": s.get("parameter10") or "",
                "parameter11": s.get("parameter11") or "",
            })

    # ── assemble workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(wb, "Policies", policy_headers,
              [{h: r.get(k, "") for h, k in zip(policy_headers, policy_keys)} for r in policy_rows])
    add_sheet(wb, "Packages", pkg_headers,
              [{h: r.get(k, "") for h, k in zip(pkg_headers, pkg_keys)} for r in pkg_rows])
    add_sheet(wb, "Scripts",  script_headers,
              [{h: r.get(k, "") for h, k in zip(script_headers, script_keys)} for r in script_rows])

    # ── save ─────────────────────────────────────────────────────────────────
    import os
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    if output_file:
        base, ext = os.path.splitext(output_file)
        if ext.lower() != ".xlsx":
            ext = ".xlsx"
        out_path = f"{base}_{ts}{ext}"
    else:
        out_path = f"policies_{ts}.xlsx"

    wb.save(out_path)
    print(
        f"[✓] Written {len(policy_rows)} policies, "
        f"{len(pkg_rows)} packages, "
        f"{len(script_rows)} scripts → {out_path}",
        file=sys.stderr,
    )

# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export JAMF policies (payloads + scope) via the Classic API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # All policies as JSON (stdout)
  python jamf_policy_export.py \\
      --url https://acme.jamfcloud.com \\
      --client-id <id> --client-secret <secret>

  # Save JSON
  python jamf_policy_export.py ... --output policies.json

  # Save CSV (3 files: policies, packages, scripts)
  python jamf_policy_export.py ... --format csv --output policies.csv

  # Save as Excel workbook (3 tabs: Policies, Packages, Scripts)
  python jamf_policy_export.py ... --format xlsx --output policies.xlsx

  # Human-readable text summary
  python jamf_policy_export.py ... --format text --output policies.txt

  # Single policy by ID
  python jamf_policy_export.py ... --policy-id 42

  # Skip disabled policies
  python jamf_policy_export.py ... --no-disabled
""",
    )

    parser.add_argument("--url", required=True, metavar="JAMF_URL",
                        help="Base URL, e.g. https://acme.jamfcloud.com")
    parser.add_argument("--client-id", required=True, metavar="CLIENT_ID",
                        help="OAuth client ID (API Role client)")
    parser.add_argument("--client-secret", required=True, metavar="CLIENT_SECRET",
                        help="OAuth client secret")
    parser.add_argument("--policy-id", type=int, default=None, metavar="ID",
                        help="Fetch a single policy by ID instead of all")
    parser.add_argument("--format", choices=["json", "text", "csv", "xlsx"], default="json",
                        help="Output format: json (default), text, csv, or xlsx")
    parser.add_argument("--output", default=None, metavar="FILE",
                        help="Write to FILE instead of stdout")
    parser.add_argument("--no-disabled", action="store_true",
                        help="Skip disabled policies")
    parser.add_argument("--debug-policy", type=int, default=None, metavar="ID",
                        help="Dump the raw JSON for one policy ID and exit (for troubleshooting)")

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    jamf_url = args.url.rstrip("/")

    print("[*] Authenticating …", file=sys.stderr)
    try:
        tm = TokenManager(jamf_url, args.client_id, args.client_secret)
        _ = tm.token   # trigger initial fetch so auth errors surface immediately
    except requests.HTTPError as exc:
        print(f"[!] Authentication failed: {exc}", file=sys.stderr)
        sys.exit(1)

    # --debug-policy: dump raw + extracted record for one policy and exit
    if args.debug_policy:
        pid = args.debug_policy
        print(f"[*] Fetching raw JSON for policy id={pid} …", file=sys.stderr)
        raw = classic_get(f"{jamf_url}/JSSResource/policies/id/{pid}", tm)
        policy = raw.get("policy", raw)   # handle both wrapped and unwrapped
        print("\n" + "=" * 60)
        print("RAW API RESPONSE (policy key):")
        print("=" * 60)
        print(json.dumps(policy, indent=2, default=str))
        print("\n" + "=" * 60)
        print("EXTRACTED RECORD:")
        print("=" * 60)
        extracted = build_policy_record(policy)
        print(json.dumps(extracted, indent=2, default=str))
        print("\n" + "=" * 60)
        print("EXTRACTION SUMMARY:")
        print("=" * 60)
        p = extracted["payloads"]
        print(f"  packages   : {len(p['packages'])} → {[x['name'] for x in p['packages']]}")
        print(f"  scripts    : {len(p['scripts'])} → {[x['name'] for x in p['scripts']]}")
        print(f"  printers   : {len(p['printers'])}")
        print(f"  dock_items : {len(p['dock_items'])}")
        print(f"  scope all_computers: {extracted['scope']['all_computers']}")
        print(f"  scope computers   : {[x.get('name') for x in extracted['scope']['computers']]}")
        print(f"  scope groups      : {[x.get('name') for x in extracted['scope']['computer_groups']]}")
        sys.exit(0)

    # Collect policy IDs to process
    if args.policy_id:
        policy_stubs = [{"id": args.policy_id, "name": "(single lookup)"}]
    else:
        print("[*] Fetching policy list …", file=sys.stderr)
        policy_stubs = get_all_policy_ids(jamf_url, tm)
        print(f"[*] Found {len(policy_stubs)} policies", file=sys.stderr)

    # Fetch details — TokenManager handles mid-run 401s automatically
    records = []
    errors = []
    for stub in policy_stubs:
        pid = stub["id"]
        try:
            detail = get_policy_detail(jamf_url, tm, pid)
            record = build_policy_record(detail)

            if args.no_disabled and not record["enabled"]:
                continue

            records.append(record)
            p = record["payloads"]
            pkg_count = len(p["packages"])
            scr_count = len(p["scripts"])
            payload_hint = []
            if pkg_count: payload_hint.append(f"{pkg_count} pkg(s)")
            if scr_count: payload_hint.append(f"{scr_count} script(s)")
            hint_str = f"  [{', '.join(payload_hint)}]" if payload_hint else ""
            print(f"  [{len(records):>4}] {record['name']} (id={pid}){hint_str}", file=sys.stderr)

        except requests.HTTPError as exc:
            msg = f"Policy id={pid} — HTTP {exc.response.status_code}: {exc}"
            print(f"  [!] {msg}", file=sys.stderr)
            errors.append(msg)
        except Exception as exc:  # noqa: BLE001
            msg = f"Policy id={pid} — unexpected error: {exc}"
            print(f"  [!] {msg}", file=sys.stderr)
            errors.append(msg)

    print(f"[*] Processed {len(records)} policies ({len(errors)} errors)", file=sys.stderr)

    # Write output
    if args.format == "json":
        write_json(records, args.output)
    elif args.format == "csv":
        write_csv(records, args.output)
    elif args.format == "xlsx":
        write_xlsx(records, args.output)
    else:
        write_text_summary(records, args.output)

    if errors:
        print("\n[!] Errors encountered:", file=sys.stderr)
        for e in errors:
            print(f"    {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
