#!/usr/bin/env python3
"""
JAMF Classic API - macOS Configuration Profile Export Script
Fetches all OS X Configuration Profiles with full detail (general info,
scope, self service, and the embedded payload XML) using OAuth client
credentials.

Token is automatically refreshed when near expiry or on 401.
Output formats: json, text, csv (3 files: profiles, scope, payloads).
"""

import argparse
import csv
import datetime
import io
import json
import sys
import time
import xml.etree.ElementTree as ET
import requests
from typing import Optional


# ---------------------------------------------------------------------------
# Auth — supports both OAuth (client_id/secret) and Basic (username/password)
# ---------------------------------------------------------------------------

_TOKEN_REFRESH_FRACTION   = 0.10  # refresh in last 10% of token lifetime
_TOKEN_REFRESH_FLOOR_SECS = 5     # hard floor regardless of fraction


class TokenManager:
    """
    OAuth 2.0 client-credentials token manager.
    Proactively refreshes when the token is in the last _TOKEN_REFRESH_FRACTION
    of its lifetime, and reactively refreshes on 401.
    """

    def __init__(self, jamf_url: str, client_id: str, client_secret: str) -> None:
        self._jamf_url      = jamf_url
        self._client_id     = client_id
        self._client_secret = client_secret
        self._token: Optional[str] = None
        self._expires_at: float = 0.0
        self._expires_in: int   = 0

    def _refresh_threshold(self) -> float:
        return max(_TOKEN_REFRESH_FLOOR_SECS,
                   self._expires_in * _TOKEN_REFRESH_FRACTION)

    def _fetch(self, reason: str = "initial") -> None:
        resp = requests.post(
            f"{self._jamf_url}/api/oauth/token",
            data={"grant_type": "client_credentials",
                  "client_id": self._client_id,
                  "client_secret": self._client_secret},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        self._token      = data["access_token"]
        self._expires_in = int(data.get("expires_in", 1800))
        self._expires_at = time.monotonic() + self._expires_in
        print(
            f"[auth] OAuth token acquired ({reason}) — valid for {self._expires_in}s, "
            f"will refresh with ~{self._refresh_threshold():.0f}s remaining",
            file=sys.stderr,
        )

    @property
    def token(self) -> str:
        remaining = self._expires_at - time.monotonic()
        if self._token is None or remaining < self._refresh_threshold():
            if self._token is not None:
                print(f"[auth] Token expiring soon ({remaining:.0f}s left) — refreshing …",
                      file=sys.stderr)
                self._fetch(reason="proactive refresh")
            else:
                self._fetch(reason="initial")
        return self._token  # type: ignore[return-value]

    def force_refresh(self) -> None:
        print("[auth] 401 received — forcing token refresh …", file=sys.stderr)
        self._token      = None
        self._expires_at = 0.0
        self._fetch(reason="forced after 401")


class BasicAuthSession:
    """
    Username/password auth for the Classic API.
    Exchanges credentials for a short-lived Bearer token via
    /api/v1/auth/token on first use, then refreshes proactively
    (same fraction logic as TokenManager) or reactively on 401.

    Falls back to HTTP Basic Auth on the request if the token
    endpoint is unavailable, so it also works on older JAMF versions.
    """

    def __init__(self, jamf_url: str, username: str, password: str) -> None:
        self._jamf_url  = jamf_url
        self._username  = username
        self._password  = password
        self._token: Optional[str] = None
        self._expires_at: float = 0.0
        self._expires_in: int   = 1800   # assume 30 min until we know better
        self._use_basic = False           # fallback flag

    def _refresh_threshold(self) -> float:
        return max(_TOKEN_REFRESH_FLOOR_SECS,
                   self._expires_in * _TOKEN_REFRESH_FRACTION)

    def _fetch(self, reason: str = "initial") -> None:
        """Exchange username/password for a Bearer token via /api/v1/auth/token."""
        resp = requests.post(
            f"{self._jamf_url}/api/v1/auth/token",
            auth=(self._username, self._password),
            timeout=30,
        )
        if resp.status_code in (404, 405):
            # Older JAMF Pro — token endpoint doesn't exist; fall back to Basic Auth
            print("[auth] /api/v1/auth/token not available — using HTTP Basic Auth",
                  file=sys.stderr)
            self._use_basic = True
            self._token      = None
            self._expires_at = time.monotonic() + 315360000  # effectively never expires
            return

        resp.raise_for_status()
        data = resp.json()
        self._token = data.get("token") or data.get("access_token")
        # JAMF returns expires in ISO 8601; derive seconds from now
        import dateutil.parser as dp  # type: ignore
        try:
            expires_str = data.get("expires") or data.get("expires_at") or ""
            if expires_str:
                import datetime as _dt
                exp = dp.parse(expires_str)
                now = _dt.datetime.now(_dt.timezone.utc)
                self._expires_in = max(60, int((exp - now).total_seconds()))
            else:
                self._expires_in = 1800
        except Exception:
            self._expires_in = 1800
        self._expires_at = time.monotonic() + self._expires_in
        print(
            f"[auth] Basic→Bearer token acquired ({reason}) — "
            f"valid ~{self._expires_in}s, "
            f"refresh with ~{self._refresh_threshold():.0f}s remaining",
            file=sys.stderr,
        )

    @property
    def token(self) -> Optional[str]:
        """Return current token (None if using raw Basic Auth fallback)."""
        if self._use_basic:
            return None
        remaining = self._expires_at - time.monotonic()
        if self._token is None or remaining < self._refresh_threshold():
            if self._token is not None:
                print(f"[auth] Token expiring soon ({remaining:.0f}s left) — refreshing …",
                      file=sys.stderr)
                self._fetch(reason="proactive refresh")
            else:
                self._fetch(reason="initial")
        return self._token

    def force_refresh(self) -> None:
        if self._use_basic:
            return   # Basic Auth has no token to refresh
        print("[auth] 401 received — forcing token refresh …", file=sys.stderr)
        self._token      = None
        self._expires_at = 0.0
        self._fetch(reason="forced after 401")

    @property
    def basic_creds(self) -> Optional[tuple]:
        """Return (username, password) tuple when in Basic Auth fallback mode."""
        return (self._username, self._password) if self._use_basic else None


# Type alias so the rest of the code accepts either auth object
AuthSession = object   # TokenManager | BasicAuthSession


# ---------------------------------------------------------------------------
# Classic API helpers
# ---------------------------------------------------------------------------

def classic_get(url: str, auth: AuthSession, retries: int = 2) -> dict:
    """
    GET a Classic API endpoint and return parsed JSON.
    Accepts either a TokenManager (OAuth) or BasicAuthSession (username/password).
    Automatically refreshes credentials and retries up to `retries` times on 401.
    """
    for attempt in range(1, retries + 2):
        headers = {"Accept": "application/json"}

        if isinstance(auth, BasicAuthSession) and auth.basic_creds:
            # Raw Basic Auth fallback (older JAMF)
            resp = requests.get(url, headers=headers,
                                auth=auth.basic_creds, timeout=60)
        else:
            # Bearer token (OAuth or Basic→Bearer)
            tok = auth.token
            headers["Authorization"] = f"Bearer {tok}"
            resp = requests.get(url, headers=headers, timeout=60)

        if resp.status_code == 401 and attempt <= retries:
            auth.force_refresh()
            continue
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError("classic_get: exhausted retries")


def get_all_profile_stubs(jamf_url: str, auth: AuthSession) -> list[dict]:
    """Return [{id, name}, …] for every OS X configuration profile."""
    data = classic_get(f"{jamf_url}/JSSResource/osxconfigurationprofiles", auth)
    return _coerce_list(data.get("os_x_configuration_profiles", []), inner_key="os_x_configuration_profile")


def get_profile_detail(jamf_url: str, auth: AuthSession, profile_id: int) -> dict:
    """Return the full detail record for a single profile."""
    data = classic_get(f"{jamf_url}/JSSResource/osxconfigurationprofiles/id/{profile_id}", auth)
    return data.get("os_x_configuration_profile", {})


# ---------------------------------------------------------------------------
# Shape normaliser (same logic proven in policy script)
# ---------------------------------------------------------------------------

def _coerce_list(val, inner_key: str = None) -> list:
    """
    Normalise any JAMF Classic API collection field into a clean list of dicts.
    Handles: bare list, wrapped dict {"inner_key": [...]}, size-only sentinel,
    single item dict, empty list/dict, list-with-empty-strings.
    """
    if not val:
        return []
    if isinstance(val, list):
        return [i for i in val if isinstance(i, dict) and i]
    if isinstance(val, dict):
        without_size = {k: v for k, v in val.items() if k != "size"}
        if not without_size:
            return []
        if inner_key and inner_key in val:
            inner = val[inner_key]
            if isinstance(inner, list):
                return [i for i in inner if isinstance(i, dict) and i]
            if isinstance(inner, dict) and inner:
                return [inner]
            return []
        return [without_size]
    return []


def _safe_get(obj, *keys, default=None):
    cur = obj
    for key in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key, default)
        if cur is None:
            return default
    return cur


# ---------------------------------------------------------------------------
# Payload XML parser
# ---------------------------------------------------------------------------

def parse_payloads_xml(xml_str: str) -> list[dict]:
    """
    Parse the mobileconfig payload XML embedded in the profile record.
    Returns a list of payload dicts, one per PayloadContent item, with:
      - PayloadType, PayloadDisplayName, PayloadDescription,
        PayloadIdentifier, PayloadUUID, PayloadVersion
      - all remaining keys flattened as key→value pairs (strings)

    The XML is an Apple plist; we extract the PayloadContent array.
    Returns [] if xml_str is empty/None or fails to parse.
    """
    if not xml_str or not xml_str.strip():
        return []

    try:
        root = ET.fromstring(xml_str.strip())
    except ET.ParseError:
        return [{"parse_error": "could not parse payload XML"}]

    # Plist structure: <plist><dict>...<key>PayloadContent</key><array>...</array>...</dict></plist>
    # Walk to find PayloadContent array
    def plist_dict_to_python(elem) -> dict:
        """Convert a plist <dict> element to a Python dict (one level deep)."""
        result = {}
        children = list(elem)
        i = 0
        while i < len(children) - 1:
            if children[i].tag == "key":
                key = children[i].text or ""
                val_elem = children[i + 1]
                tag = val_elem.tag
                if tag == "string":
                    result[key] = val_elem.text or ""
                elif tag == "integer":
                    result[key] = val_elem.text or ""
                elif tag == "real":
                    result[key] = val_elem.text or ""
                elif tag == "true":
                    result[key] = "true"
                elif tag == "false":
                    result[key] = "false"
                elif tag == "data":
                    result[key] = "<data>"
                elif tag == "array":
                    result[key] = f"<array[{len(list(val_elem))}]>"
                elif tag == "dict":
                    result[key] = "<dict>"
                else:
                    result[key] = f"<{tag}>"
                i += 2
            else:
                i += 1
        return result

    # Find the top-level dict
    top_dict = root.find("dict")
    if top_dict is None:
        # Some profiles wrap in <plist version="1.0"><dict>
        top_dict = root if root.tag == "dict" else None
    if top_dict is None:
        return [{"parse_error": "no top-level dict found"}]

    # Find PayloadContent array
    children = list(top_dict)
    payload_array = None
    for i, child in enumerate(children):
        if child.tag == "key" and child.text == "PayloadContent":
            if i + 1 < len(children) and children[i + 1].tag == "array":
                payload_array = children[i + 1]
            break

    if payload_array is None:
        # No PayloadContent — return the top-level dict keys as a single payload
        return [plist_dict_to_python(top_dict)]

    payloads = []
    for item in payload_array:
        if item.tag == "dict":
            payloads.append(plist_dict_to_python(item))
    return payloads


# ---------------------------------------------------------------------------
# Scope extraction (same _coerce_list approach as policy script)
# ---------------------------------------------------------------------------

def extract_scope(profile: dict) -> dict:
    scope = profile.get("scope", {}) or {}

    def grab(container, key: str, inner_key: str) -> list:
        if not isinstance(container, dict):
            return []
        return _coerce_list(container.get(key, []), inner_key=inner_key)

    lim  = scope.get("limitations") or {}
    excl = scope.get("exclusions")  or {}

    return {
        "all_computers":   scope.get("all_computers", False),
        "all_jss_users":   scope.get("all_jss_users", False),
        "computers":       grab(scope, "computers",       "computer"),
        "computer_groups": grab(scope, "computer_groups", "computer_group"),
        "buildings":       grab(scope, "buildings",       "building"),
        "departments":     grab(scope, "departments",     "department"),
        "jss_users":       grab(scope, "jss_users",       "jss_user"),
        "jss_user_groups": grab(scope, "jss_user_groups", "jss_user_group"),
        "limitations": {
            "users":            grab(lim, "users",            "user"),
            "user_groups":      grab(lim, "user_groups",      "user_group"),
            "network_segments": grab(lim, "network_segments", "network_segment"),
            "ibeacons":         grab(lim, "ibeacons",         "ibeacon"),
        },
        "exclusions": {
            "computers":        grab(excl, "computers",        "computer"),
            "computer_groups":  grab(excl, "computer_groups",  "computer_group"),
            "buildings":        grab(excl, "buildings",        "building"),
            "departments":      grab(excl, "departments",      "department"),
            "users":            grab(excl, "users",            "user"),
            "user_groups":      grab(excl, "user_groups",      "user_group"),
            "network_segments": grab(excl, "network_segments", "network_segment"),
            "ibeacons":         grab(excl, "ibeacons",         "ibeacon"),
        },
    }


# ---------------------------------------------------------------------------
# Self service extraction
# ---------------------------------------------------------------------------

def extract_self_service(profile: dict) -> dict:
    ss = profile.get("self_service", {}) or {}
    return {
        "use_for_self_service":          ss.get("use_for_self_service", False),
        "self_service_display_name":     ss.get("self_service_display_name", ""),
        "install_button_text":           ss.get("install_button_text", ""),
        "self_service_description":      ss.get("self_service_description", ""),
        "force_users_to_view_description": ss.get("force_users_to_view_description", False),
        "feature_on_main_page":          ss.get("feature_on_main_page", False),
        "notification":                  ss.get("notification", False),
        "notification_subject":          ss.get("notification_subject", ""),
        "notification_message":          ss.get("notification_message", ""),
        "self_service_categories":       _coerce_list(ss.get("self_service_categories", []),
                                                      inner_key="category"),
    }


# ---------------------------------------------------------------------------
# Main record builder
# ---------------------------------------------------------------------------

def build_profile_record(profile: dict) -> dict:
    general = profile.get("general", {}) or {}

    # Parse the embedded mobileconfig payload XML
    payloads_xml = general.get("payloads", "") or ""
    parsed_payloads = parse_payloads_xml(payloads_xml)

    return {
        "id":          general.get("id"),
        "name":        general.get("name"),
        "description": general.get("description", ""),
        "site":        _safe_get(general, "site", "name"),
        "category":    _safe_get(general, "category", "name"),
        "distribution_method": general.get("distribution_method", ""),
        "user_removable":      general.get("user_removable", False),
        "level":               general.get("level", ""),   # "Computer" or "User"
        "uuid":                general.get("uuid", ""),
        "redeploy_on_update":  general.get("redeploy_on_update", ""),
        "payloads_xml":    payloads_xml,       # raw mobileconfig XML
        "payloads_parsed": parsed_payloads,    # list of payload dicts
        "scope":         extract_scope(profile),
        "self_service":  extract_self_service(profile),
    }


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def _names(items: list) -> str:
    return "; ".join(i.get("name", "") for i in items if isinstance(i, dict) and i.get("name"))


def _scope_summary(sc: dict) -> str:
    if sc.get("all_computers"):
        return "All Computers"
    parts = []
    for c  in sc.get("computers", []):       parts.append(c.get("name", ""))
    for g  in sc.get("computer_groups", []): parts.append(f"[Group] {g.get('name', '')}")
    for b  in sc.get("buildings", []):       parts.append(f"[Building] {b.get('name', '')}")
    for d  in sc.get("departments", []):     parts.append(f"[Dept] {d.get('name', '')}")
    return "; ".join(p for p in parts if p)


def _stamp(path: Optional[str]) -> Optional[str]:
    """Insert YYYYMMDD_HHMMSS timestamp before the file extension."""
    if not path:
        return path
    import os
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    return f"{base}_{ts}{ext}"


# ---------------------------------------------------------------------------
# JSON output
# ---------------------------------------------------------------------------

def write_json(records: list[dict], output_file: Optional[str]) -> None:
    # Exclude raw XML from JSON output to keep it readable (parsed version is there)
    clean = []
    for r in records:
        row = dict(r)
        row.pop("payloads_xml", None)
        clean.append(row)
    out = json.dumps(clean, indent=2, default=str)
    output_file = _stamp(output_file)
    if output_file:
        with open(output_file, "w") as f:
            f.write(out)
        print(f"[✓] Written {len(records)} profiles → {output_file}", file=sys.stderr)
    else:
        print(out)


# ---------------------------------------------------------------------------
# Text summary output
# ---------------------------------------------------------------------------

def write_text_summary(records: list[dict], output_file: Optional[str]) -> None:
    lines = []
    sep = "=" * 72

    for rec in records:
        lines.append(sep)
        lines.append(f"Profile: {rec['name']}  (ID: {rec['id']})")
        lines.append(f"  Description       : {rec['description']}")
        lines.append(f"  Category          : {rec['category']}")
        lines.append(f"  Level             : {rec['level']}")
        lines.append(f"  Distribution      : {rec['distribution_method']}")
        lines.append(f"  User Removable    : {rec['user_removable']}")
        lines.append(f"  Redeploy on Update: {rec['redeploy_on_update']}")

        # Parsed payloads
        pp = rec["payloads_parsed"]
        if pp:
            lines.append(f"  Payloads ({len(pp)}):")
            for pl in pp:
                ptype = pl.get("PayloadType", pl.get("PayloadDisplayName", "unknown"))
                pname = pl.get("PayloadDisplayName", "")
                pdesc = pl.get("PayloadDescription", "")
                lines.append(f"    • {ptype}" + (f" — {pname}" if pname and pname != ptype else ""))
                if pdesc:
                    lines.append(f"        {pdesc}")
        else:
            lines.append("  Payloads: (none / not parseable)")

        # Scope
        sc = rec["scope"]
        lines.append("  Scope:")
        if sc["all_computers"]:
            lines.append("    Targets: All Computers")
        else:
            for c  in sc["computers"]:       lines.append(f"    Computer      : {c.get('name')} (id={c.get('id')})")
            for g  in sc["computer_groups"]: lines.append(f"    Computer Group: {g.get('name')} (id={g.get('id')})")
            for b  in sc["buildings"]:       lines.append(f"    Building      : {b.get('name')}")
            for d  in sc["departments"]:     lines.append(f"    Department    : {d.get('name')}")

        excl = sc["exclusions"]
        excl_items = (excl["computers"] + excl["computer_groups"] +
                      excl["buildings"] + excl["departments"])
        if excl_items:
            lines.append("    Exclusions:")
            for item in excl_items:
                lines.append(f"      - {item.get('name')}")

        lim = sc["limitations"]
        lim_items = lim["users"] + lim["user_groups"] + lim["network_segments"]
        if lim_items:
            lines.append("    Limitations:")
            for item in lim_items:
                lines.append(f"      - {item.get('name')}")

        # Self service
        ss = rec["self_service"]
        if ss["use_for_self_service"]:
            lines.append(f"  Self Service: {ss['self_service_display_name'] or rec['name']}")
            if ss["self_service_description"]:
                lines.append(f"    Description: {ss['self_service_description']}")

    lines.append(sep)
    output = "\n".join(lines)
    output_file = _stamp(output_file)
    if output_file:
        with open(output_file, "w") as f:
            f.write(output)
        print(f"[✓] Written {len(records)} profiles → {output_file}", file=sys.stderr)
    else:
        print(output)


# ---------------------------------------------------------------------------
# CSV output — 3 files: profiles, scope, payloads
# ---------------------------------------------------------------------------

def write_csv(records: list[dict], output_file: Optional[str]) -> None:
    """
    Writes three CSV files:
      profiles_TIMESTAMP.csv       — one row per profile (general + scope summary)
      profiles_scope_TIMESTAMP.csv — one row per scope target/exclusion/limitation
      profiles_payloads_TIMESTAMP.csv — one row per parsed payload within a profile
    """

    # --- profiles.csv ---
    profile_fields = [
        "profile_id", "profile_name", "description", "category", "site",
        "level", "distribution_method", "user_removable", "redeploy_on_update", "uuid",
        "payload_count",
        "payload_types",          # semicolon list of PayloadType values
        # Scope summary
        "scope_all_computers", "scope_targets",
        "scope_computers", "scope_computer_groups",
        "scope_buildings", "scope_departments",
        "scope_limit_users", "scope_limit_user_groups", "scope_limit_network_segments",
        "scope_excl_computers", "scope_excl_groups",
        "scope_excl_buildings", "scope_excl_departments",
        # Self service
        "self_service_enabled", "self_service_name",
        "self_service_description", "self_service_button_text",
        "self_service_notification",
    ]

    # --- scope.csv ---
    # One row per target/exclusion/limitation entry
    scope_fields = [
        "profile_id", "profile_name", "scope_type",   # "target" / "exclusion" / "limitation"
        "entry_kind",  # "computer" / "computer_group" / "building" / "department" / "user" / etc.
        "entry_id", "entry_name",
    ]

    # --- payloads.csv ---
    # One row per parsed payload block inside the mobileconfig
    payload_fields = [
        "profile_id", "profile_name", "category",
        "payload_index",
        "PayloadType", "PayloadDisplayName", "PayloadDescription",
        "PayloadIdentifier", "PayloadUUID", "PayloadVersion",
        "extra_keys",   # JSON of any remaining keys not in the standard set above
    ]

    STANDARD_PAYLOAD_KEYS = {
        "PayloadType", "PayloadDisplayName", "PayloadDescription",
        "PayloadIdentifier", "PayloadUUID", "PayloadVersion",
    }

    profile_rows, scope_rows, payload_rows = [], [], []

    for rec in records:
        pid   = rec["id"]
        pname = rec["name"]
        sc    = rec["scope"]
        ss    = rec["self_service"]
        pp    = rec["payloads_parsed"]

        payload_types = "; ".join(
            p.get("PayloadType", "") for p in pp if p.get("PayloadType")
        )

        profile_rows.append({
            "profile_id":   pid,
            "profile_name": pname,
            "description":  rec["description"],
            "category":     rec["category"],
            "site":         rec["site"],
            "level":        rec["level"],
            "distribution_method": rec["distribution_method"],
            "user_removable":      rec["user_removable"],
            "redeploy_on_update":  rec["redeploy_on_update"],
            "uuid":                rec["uuid"],
            "payload_count":       len(pp),
            "payload_types":       payload_types,
            "scope_all_computers": sc.get("all_computers"),
            "scope_targets":       _scope_summary(sc),
            "scope_computers":     _names(sc.get("computers", [])),
            "scope_computer_groups": _names(sc.get("computer_groups", [])),
            "scope_buildings":     _names(sc.get("buildings", [])),
            "scope_departments":   _names(sc.get("departments", [])),
            "scope_limit_users":   _names(sc.get("limitations", {}).get("users", [])),
            "scope_limit_user_groups": _names(sc.get("limitations", {}).get("user_groups", [])),
            "scope_limit_network_segments": _names(sc.get("limitations", {}).get("network_segments", [])),
            "scope_excl_computers":  _names(sc.get("exclusions", {}).get("computers", [])),
            "scope_excl_groups":     _names(sc.get("exclusions", {}).get("computer_groups", [])),
            "scope_excl_buildings":  _names(sc.get("exclusions", {}).get("buildings", [])),
            "scope_excl_departments":_names(sc.get("exclusions", {}).get("departments", [])),
            "self_service_enabled":      ss.get("use_for_self_service"),
            "self_service_name":         ss.get("self_service_display_name"),
            "self_service_description":  ss.get("self_service_description"),
            "self_service_button_text":  ss.get("install_button_text"),
            "self_service_notification": ss.get("notification"),
        })

        # Scope detail rows
        def add_scope_rows(items, kind, stype):
            for item in items:
                if not isinstance(item, dict):
                    continue
                scope_rows.append({
                    "profile_id":   pid,
                    "profile_name": pname,
                    "scope_type":   stype,
                    "entry_kind":   kind,
                    "entry_id":     item.get("id", ""),
                    "entry_name":   item.get("name", ""),
                })

        add_scope_rows(sc.get("computers", []),       "computer",       "target")
        add_scope_rows(sc.get("computer_groups", []), "computer_group", "target")
        add_scope_rows(sc.get("buildings", []),       "building",       "target")
        add_scope_rows(sc.get("departments", []),     "department",     "target")
        add_scope_rows(sc.get("exclusions", {}).get("computers", []),       "computer",       "exclusion")
        add_scope_rows(sc.get("exclusions", {}).get("computer_groups", []), "computer_group", "exclusion")
        add_scope_rows(sc.get("exclusions", {}).get("buildings", []),       "building",       "exclusion")
        add_scope_rows(sc.get("exclusions", {}).get("departments", []),     "department",     "exclusion")
        add_scope_rows(sc.get("limitations", {}).get("users", []),            "user",            "limitation")
        add_scope_rows(sc.get("limitations", {}).get("user_groups", []),      "user_group",      "limitation")
        add_scope_rows(sc.get("limitations", {}).get("network_segments", []), "network_segment", "limitation")

        # Payload rows
        for idx, pl in enumerate(pp):
            extra = {k: v for k, v in pl.items() if k not in STANDARD_PAYLOAD_KEYS}
            payload_rows.append({
                "profile_id":          pid,
                "profile_name":        pname,
                "category":            rec["category"],
                "payload_index":       idx + 1,
                "PayloadType":         pl.get("PayloadType", ""),
                "PayloadDisplayName":  pl.get("PayloadDisplayName", ""),
                "PayloadDescription":  pl.get("PayloadDescription", ""),
                "PayloadIdentifier":   pl.get("PayloadIdentifier", ""),
                "PayloadUUID":         pl.get("PayloadUUID", ""),
                "PayloadVersion":      pl.get("PayloadVersion", ""),
                "extra_keys":          json.dumps(extra, default=str) if extra else "",
            })

    def _write(rows, fields, path, label):
        if path:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=fields)
                w.writeheader()
                w.writerows(rows)
            print(f"[✓] Written {len(rows)} rows → {path}  ({label})", file=sys.stderr)
        else:
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=fields)
            w.writeheader()
            w.writerows(rows)
            print(f"\n### {label} ###")
            print(buf.getvalue())

    # Build timestamped output paths
    if output_file:
        import os
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_file)
        output_file  = f"{base}_{ts}{ext}"
        scope_path   = f"{base}_scope_{ts}{ext}"
        payload_path = f"{base}_payloads_{ts}{ext}"
    else:
        scope_path = payload_path = None

    _write(profile_rows, profile_fields, output_file,  "profiles")
    _write(scope_rows,   scope_fields,   scope_path,   "scope")
    _write(payload_rows, payload_fields, payload_path, "payloads")




# ---------------------------------------------------------------------------
# XLSX workbook output — one tab per data type, formatted
# ---------------------------------------------------------------------------

def write_xlsx(records: list[dict], output_file: Optional[str]) -> None:
    """
    Write a single timestamped .xlsx workbook with four tabs:
      • Profiles   — one row per profile (general + scope summary + self service)
      • Scope      — one row per scope/exclusion/limitation entry
      • Payloads   — one row per parsed mobileconfig payload block
      • Raw XML    — one row per profile with the full mobileconfig XML
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── colour palette ──────────────────────────────────────────────────────
    HDR_BG   = "1F4E79"   # dark navy
    HDR_FG   = "FFFFFF"   # white text
    ALT_BG   = "D6E4F0"   # light blue alternating row
    TAB_COLS = {           # tab accent colours
        "Profiles": "1F4E79",
        "Scope":    "2E75B6",
        "Payloads": "2F5496",
        "Raw XML":  "4472C4",
    }

    thin = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers: list[str]) -> None:
        hdr_font  = Font(name="Arial", bold=True, color=HDR_FG, size=10)
        hdr_fill  = PatternFill("solid", fgColor=HDR_BG)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    def write_rows(ws, headers: list[str], rows: list[dict], alt_col: str = ALT_BG) -> None:
        alt_fill = PatternFill("solid", fgColor=alt_col)
        body_font = Font(name="Arial", size=9)
        wrap_align = Alignment(vertical="top", wrap_text=False)
        for r_idx, row in enumerate(rows, start=2):
            fill = alt_fill if r_idx % 2 == 0 else None
            for c_idx, key in enumerate(headers, start=1):
                val = row.get(key, "")
                if val is None:
                    val = ""
                elif isinstance(val, bool):
                    val = "Yes" if val else "No"
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if not isinstance(val, (int, float)) else val)
                cell.font      = body_font
                cell.alignment = wrap_align
                cell.border    = cell_border
                if fill:
                    cell.fill = fill

    def autofit(ws, headers: list[str], rows: list[dict], min_w: int = 10, max_w: int = 60) -> None:
        for c_idx, key in enumerate(headers, start=1):
            col_letter = get_column_letter(c_idx)
            # Sample up to 200 rows for width estimation
            vals = [str(key)] + [str(r.get(key, "") or "") for r in rows[:200]]
            best = min(max(max(len(v.split("\n")[0]) for v in vals) + 2, min_w), max_w)
            ws.column_dimensions[col_letter].width = best

    def add_sheet(wb, name: str, headers: list[str], rows: list[dict]) -> None:
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = TAB_COLS.get(name, "4472C4")
        style_header(ws, headers)
        write_rows(ws, headers, rows)
        autofit(ws, headers, rows)
        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── build row data (same logic as write_csv) ─────────────────────────────
    STANDARD_PAYLOAD_KEYS = {
        "PayloadType", "PayloadDisplayName", "PayloadDescription",
        "PayloadIdentifier", "PayloadUUID", "PayloadVersion",
    }

    profile_headers = [
        "Profile ID", "Profile Name", "Description", "Category", "Site",
        "Level", "Distribution Method", "User Removable", "Redeploy on Update", "UUID",
        "Payload Count", "Payload Types",
        "Scope: All Computers", "Scope: Targets",
        "Scope: Computers", "Scope: Computer Groups",
        "Scope: Buildings", "Scope: Departments",
        "Limit: Users", "Limit: User Groups", "Limit: Network Segments",
        "Excl: Computers", "Excl: Computer Groups",
        "Excl: Buildings", "Excl: Departments",
        "Self Service Enabled", "Self Service Name",
        "Self Service Description", "Self Service Button",
        "Self Service Notification",
    ]
    profile_keys = [
        "profile_id", "profile_name", "description", "category", "site",
        "level", "distribution_method", "user_removable", "redeploy_on_update", "uuid",
        "payload_count", "payload_types",
        "scope_all_computers", "scope_targets",
        "scope_computers", "scope_computer_groups",
        "scope_buildings", "scope_departments",
        "scope_limit_users", "scope_limit_user_groups", "scope_limit_network_segments",
        "scope_excl_computers", "scope_excl_groups",
        "scope_excl_buildings", "scope_excl_departments",
        "self_service_enabled", "self_service_name",
        "self_service_description", "self_service_button_text",
        "self_service_notification",
    ]

    scope_headers = [
        "Profile ID", "Profile Name",
        "Scope Type", "Entry Kind", "Entry ID", "Entry Name",
    ]
    scope_keys = [
        "profile_id", "profile_name",
        "scope_type", "entry_kind", "entry_id", "entry_name",
    ]

    payload_headers = [
        "Profile ID", "Profile Name", "Category", "Payload #",
        "Payload Type", "Display Name", "Description",
        "Identifier", "UUID", "Version", "Extra Keys",
    ]
    payload_keys = [
        "profile_id", "profile_name", "category", "payload_index",
        "PayloadType", "PayloadDisplayName", "PayloadDescription",
        "PayloadIdentifier", "PayloadUUID", "PayloadVersion", "extra_keys",
    ]

    xml_headers = ["Profile ID", "Profile Name", "Category", "Raw Mobileconfig XML"]
    xml_keys    = ["profile_id", "profile_name", "category", "payloads_xml"]

    profile_rows, scope_rows, payload_rows, xml_rows = [], [], [], []

    for rec in records:
        pid   = rec["id"]
        pname = rec["name"]
        sc    = rec["scope"]
        ss    = rec["self_service"]
        pp    = rec["payloads_parsed"]

        payload_types = "; ".join(p.get("PayloadType", "") for p in pp if p.get("PayloadType"))

        profile_rows.append({
            "profile_id": pid, "profile_name": pname,
            "description": rec["description"], "category": rec["category"],
            "site": rec["site"], "level": rec["level"],
            "distribution_method": rec["distribution_method"],
            "user_removable": rec["user_removable"],
            "redeploy_on_update": rec["redeploy_on_update"], "uuid": rec["uuid"],
            "payload_count": len(pp), "payload_types": payload_types,
            "scope_all_computers": sc.get("all_computers"),
            "scope_targets": _scope_summary(sc),
            "scope_computers": _names(sc.get("computers", [])),
            "scope_computer_groups": _names(sc.get("computer_groups", [])),
            "scope_buildings": _names(sc.get("buildings", [])),
            "scope_departments": _names(sc.get("departments", [])),
            "scope_limit_users": _names(sc.get("limitations", {}).get("users", [])),
            "scope_limit_user_groups": _names(sc.get("limitations", {}).get("user_groups", [])),
            "scope_limit_network_segments": _names(sc.get("limitations", {}).get("network_segments", [])),
            "scope_excl_computers": _names(sc.get("exclusions", {}).get("computers", [])),
            "scope_excl_groups": _names(sc.get("exclusions", {}).get("computer_groups", [])),
            "scope_excl_buildings": _names(sc.get("exclusions", {}).get("buildings", [])),
            "scope_excl_departments": _names(sc.get("exclusions", {}).get("departments", [])),
            "self_service_enabled": ss.get("use_for_self_service"),
            "self_service_name": ss.get("self_service_display_name"),
            "self_service_description": ss.get("self_service_description"),
            "self_service_button_text": ss.get("install_button_text"),
            "self_service_notification": ss.get("notification"),
        })

        def add_scope(items, kind, stype):
            for item in items:
                if not isinstance(item, dict):
                    continue
                scope_rows.append({
                    "profile_id": pid, "profile_name": pname,
                    "scope_type": stype, "entry_kind": kind,
                    "entry_id": item.get("id", ""), "entry_name": item.get("name", ""),
                })
        add_scope(sc.get("computers", []),       "computer",       "target")
        add_scope(sc.get("computer_groups", []), "computer_group", "target")
        add_scope(sc.get("buildings", []),       "building",       "target")
        add_scope(sc.get("departments", []),     "department",     "target")
        add_scope(sc.get("exclusions", {}).get("computers", []),       "computer",       "exclusion")
        add_scope(sc.get("exclusions", {}).get("computer_groups", []), "computer_group", "exclusion")
        add_scope(sc.get("exclusions", {}).get("buildings", []),       "building",       "exclusion")
        add_scope(sc.get("exclusions", {}).get("departments", []),     "department",     "exclusion")
        add_scope(sc.get("limitations", {}).get("users", []),            "user",            "limitation")
        add_scope(sc.get("limitations", {}).get("user_groups", []),      "user_group",      "limitation")
        add_scope(sc.get("limitations", {}).get("network_segments", []), "network_segment", "limitation")

        for idx, pl in enumerate(pp):
            extra = {k: v for k, v in pl.items() if k not in STANDARD_PAYLOAD_KEYS}
            payload_rows.append({
                "profile_id": pid, "profile_name": pname, "category": rec["category"],
                "payload_index": idx + 1,
                "PayloadType":        pl.get("PayloadType", ""),
                "PayloadDisplayName": pl.get("PayloadDisplayName", ""),
                "PayloadDescription": pl.get("PayloadDescription", ""),
                "PayloadIdentifier":  pl.get("PayloadIdentifier", ""),
                "PayloadUUID":        pl.get("PayloadUUID", ""),
                "PayloadVersion":     pl.get("PayloadVersion", ""),
                "extra_keys": json.dumps(extra, default=str) if extra else "",
            })

        xml_rows.append({
            "profile_id": pid, "profile_name": pname,
            "category": rec["category"],
            "payloads_xml": rec.get("payloads_xml", ""),
        })

    # ── assemble workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    add_sheet(wb, "Profiles", profile_headers,
              [{h: r.get(k, "") for h, k in zip(profile_headers, profile_keys)} for r in profile_rows])
    add_sheet(wb, "Scope",    scope_headers,
              [{h: r.get(k, "") for h, k in zip(scope_headers, scope_keys)} for r in scope_rows])
    add_sheet(wb, "Payloads", payload_headers,
              [{h: r.get(k, "") for h, k in zip(payload_headers, payload_keys)} for r in payload_rows])
    add_sheet(wb, "Raw XML",  xml_headers,
              [{h: r.get(k, "") for h, k in zip(xml_headers, xml_keys)} for r in xml_rows])

    # ── save ─────────────────────────────────────────────────────────────────
    import os
    if output_file:
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_file)
        if ext.lower() != ".xlsx":
            ext = ".xlsx"
        out_path = f"{base}_{ts}{ext}"
    else:
        ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = f"profiles_{ts}.xlsx"

    wb.save(out_path)
    print(
        f"[✓] Written {len(profile_rows)} profiles, "
        f"{len(scope_rows)} scope rows, "
        f"{len(payload_rows)} payload rows → {out_path}",
        file=sys.stderr,
    )

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export JAMF macOS Configuration Profiles via the Classic API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # OAuth (client credentials)
  python jamf_profiles_export.py \\
      --url https://acme.jamfcloud.com \\
      --client-id <id> --client-secret <secret>

  # Username / password
  python jamf_profiles_export.py \\
      --url https://acme.jamfcloud.com \\
      --username admin --password secret

  # Save as JSON
  python jamf_profiles_export.py ... --output profiles.json

  # Save as CSV (3 files: profiles, scope, payloads)
  python jamf_profiles_export.py ... --format csv --output profiles.csv

  # Save as Excel workbook (4 tabs: Profiles, Scope, Payloads, Raw XML)
  python jamf_profiles_export.py ... --format xlsx --output profiles.xlsx

  # Human-readable text summary
  python jamf_profiles_export.py ... --format text --output profiles.txt

  # Single profile by ID
  python jamf_profiles_export.py ... --profile-id 42

  # Debug: dump raw + extracted data for one profile
  python jamf_profiles_export.py ... --debug-profile 42
""",
    )
    # Auth — mutually exclusive groups so exactly one method must be supplied
    auth_group = parser.add_argument_group(
        "authentication (choose one method)"
    )
    auth_group.add_argument("--client-id",     default=None, metavar="CLIENT_ID",
                            help="OAuth client ID  (use with --client-secret)")
    auth_group.add_argument("--client-secret", default=None, metavar="CLIENT_SECRET",
                            help="OAuth client secret  (use with --client-id)")
    auth_group.add_argument("--username",      default=None, metavar="USERNAME",
                            help="JAMF admin username  (use with --password)")
    auth_group.add_argument("--password",      default=None, metavar="PASSWORD",
                            help="JAMF admin password  (use with --username)")

    parser.add_argument("--url",           required=True, metavar="JAMF_URL",
                        help="Base URL, e.g. https://acme.jamfcloud.com")
    parser.add_argument("--profile-id",    type=int, default=None, metavar="ID",
                        help="Fetch a single profile by ID instead of all")
    parser.add_argument("--format",        choices=["json", "text", "csv", "xlsx"], default="json",
                        help="Output format (default: json)")
    parser.add_argument("--output",        default=None, metavar="FILE",
                        help="Write to FILE instead of stdout")
    parser.add_argument("--debug-profile", type=int, default=None, metavar="ID",
                        help="Dump raw JSON + extracted record for one ID and exit")

    args = parser.parse_args()

    # Validate: exactly one auth method must be fully supplied
    has_oauth = bool(args.client_id and args.client_secret)
    has_basic = bool(args.username and args.password)
    if has_oauth and has_basic:
        parser.error("supply either --client-id/--client-secret OR --username/--password, not both")
    if not has_oauth and not has_basic:
        parser.error(
            "authentication required — supply either:\n"
            "  OAuth : --client-id <id> --client-secret <secret>\n"
            "  Basic : --username <user> --password <pass>"
        )
    if args.client_id and not args.client_secret:
        parser.error("--client-id requires --client-secret")
    if args.client_secret and not args.client_id:
        parser.error("--client-secret requires --client-id")
    if args.username and not args.password:
        parser.error("--username requires --password")
    if args.password and not args.username:
        parser.error("--password requires --username")

    return args


def main() -> None:
    args = parse_args()
    jamf_url = args.url.rstrip("/")

    print("[*] Authenticating …", file=sys.stderr)
    try:
        if args.client_id:
            auth = TokenManager(jamf_url, args.client_id, args.client_secret)
            print("[auth] Using OAuth (client credentials)", file=sys.stderr)
        else:
            auth = BasicAuthSession(jamf_url, args.username, args.password)
            print("[auth] Using Basic Auth (username/password)", file=sys.stderr)
        _ = auth.token   # trigger initial credential exchange immediately
    except requests.HTTPError as exc:
        print(f"[!] Authentication failed: {exc}", file=sys.stderr)
        sys.exit(1)

    # --debug-profile: show raw + extracted for one profile then exit
    if args.debug_profile:
        pid = args.debug_profile
        print(f"[*] Fetching raw JSON for profile id={pid} …", file=sys.stderr)
        raw = classic_get(f"{jamf_url}/JSSResource/osxconfigurationprofiles/id/{pid}", auth)
        profile = raw.get("os_x_configuration_profile", raw)
        print("\n" + "=" * 60)
        print("RAW API RESPONSE:")
        print("=" * 60)
        print(json.dumps(profile, indent=2, default=str))
        print("\n" + "=" * 60)
        print("EXTRACTED RECORD:")
        print("=" * 60)
        extracted = build_profile_record(profile)
        # Don't dump raw XML in the extracted view
        view = dict(extracted)
        view.pop("payloads_xml", None)
        print(json.dumps(view, indent=2, default=str))
        print("\n" + "=" * 60)
        print("EXTRACTION SUMMARY:")
        print("=" * 60)
        sc = extracted["scope"]
        pp = extracted["payloads_parsed"]
        print(f"  payloads parsed : {len(pp)}")
        for pl in pp:
            print(f"    • {pl.get('PayloadType','?')} — {pl.get('PayloadDisplayName','')}")
        print(f"  scope all_computers  : {sc['all_computers']}")
        print(f"  scope computers      : {[x.get('name') for x in sc['computers']]}")
        print(f"  scope groups         : {[x.get('name') for x in sc['computer_groups']]}")
        print(f"  scope excl groups    : {[x.get('name') for x in sc['exclusions']['computer_groups']]}")
        sys.exit(0)

    # Collect profile stubs
    if args.profile_id:
        stubs = [{"id": args.profile_id, "name": "(single lookup)"}]
    else:
        print("[*] Fetching profile list …", file=sys.stderr)
        stubs = get_all_profile_stubs(jamf_url, auth)
        print(f"[*] Found {len(stubs)} profiles", file=sys.stderr)

    # Fetch detail for each profile
    records = []
    errors  = []
    for stub in stubs:
        pid = stub["id"]
        try:
            detail = get_profile_detail(jamf_url, auth, pid)
            record = build_profile_record(detail)
            records.append(record)

            pp_count = len(record["payloads_parsed"])
            hint = f"  [{pp_count} payload(s)]" if pp_count else ""
            print(f"  [{len(records):>4}] {record['name']} (id={pid}){hint}", file=sys.stderr)

        except requests.HTTPError as exc:
            msg = f"Profile id={pid} — HTTP {exc.response.status_code}: {exc}"
            print(f"  [!] {msg}", file=sys.stderr)
            errors.append(msg)
        except Exception as exc:  # noqa: BLE001
            msg = f"Profile id={pid} — unexpected error: {exc}"
            print(f"  [!] {msg}", file=sys.stderr)
            errors.append(msg)

    print(f"[*] Processed {len(records)} profiles ({len(errors)} errors)", file=sys.stderr)

    if args.format == "json":
        write_json(records, args.output)
    elif args.format == "csv":
        write_csv(records, args.output)
    elif args.format == "xlsx":
        write_xlsx(records, args.output)
    else:
        write_text_summary(records, args.output)

    if errors:
        print("\n[!] Errors encountered:", file=sys.stderr)
        for e in errors:
            print(f"    {e}", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()